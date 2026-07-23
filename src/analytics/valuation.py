"""
src/analytics/valuation.py
Valuation analysis module:
- FCF yield for all 92 companies
- Sector median P/E with overvaluation flags (Caution/Discount/Fair)
- 5-year median P/E per company
- Output: output/valuation_summary.xlsx, output/valuation_flags.csv
"""
import os
import sqlite3
import logging
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

DB_PATH = "data/nifty100.db"
ACTIVE_YEAR = "2024-03"
ACTIVE_YR_INT = 2024
HIST_YEARS = [2020, 2021, 2022, 2023, 2024]

# Cell fills
CAUTION_FILL  = PatternFill(start_color="FFB74D", end_color="FFB74D", fill_type="solid")  # orange
DISCOUNT_FILL = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")  # green
FAIR_FILL     = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # light blue
HEADER_FILL   = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)


def get_conn():
    return sqlite3.connect(DB_PATH)


def load_data() -> pd.DataFrame:
    conn = get_conn()

    df_fr = pd.read_sql(
        "SELECT company_id, free_cash_flow_cr, composite_quality_score "
        "FROM financial_ratios WHERE year = ?",
        conn, params=(ACTIVE_YEAR,)
    )
    df_c = pd.read_sql("SELECT id as company_id, company_name FROM companies", conn)
    df_s = pd.read_sql("SELECT company_id, broad_sector FROM sectors", conn)

    # Current year valuation
    df_mc_cur = pd.read_sql(
        "SELECT company_id, market_cap_crore, pe_ratio, pb_ratio, ev_ebitda, dividend_yield_pct "
        "FROM market_cap WHERE year = ?",
        conn, params=(ACTIVE_YR_INT,)
    )

    # Historical P/E for 5-year median
    placeholders = ",".join("?" * len(HIST_YEARS))
    df_mc_hist = pd.read_sql(
        f"SELECT company_id, year, pe_ratio FROM market_cap WHERE year IN ({placeholders})",
        conn, params=HIST_YEARS
    )

    conn.close()
    return df_fr, df_c, df_s, df_mc_cur, df_mc_hist


def compute_five_yr_median_pe(df_hist: pd.DataFrame) -> pd.DataFrame:
    """Compute 5-year median P/E per company."""
    medians = (df_hist.groupby("company_id")["pe_ratio"]
               .apply(lambda x: np.nanmedian(x.values) if x.notna().any() else np.nan)
               .reset_index())
    medians.columns = ["company_id", "five_yr_median_pe"]
    return medians


def compute_sector_median_pe(df: pd.DataFrame) -> pd.DataFrame:
    """Compute sector median P/E for the current year."""
    sector_pe = (df.groupby("broad_sector")["pe_ratio"]
                 .apply(lambda x: np.nanmedian(x.values) if x.notna().any() else np.nan)
                 .reset_index())
    sector_pe.columns = ["broad_sector", "sector_median_pe"]
    return sector_pe


def apply_valuation_flag(pe, sector_pe):
    """Apply Caution/Discount/Fair label."""
    if pd.isna(pe) or pd.isna(sector_pe) or sector_pe == 0:
        return "N/A"
    ratio = pe / sector_pe
    if ratio > 1.5:
        return "Caution"
    elif ratio < 0.7:
        return "Discount"
    else:
        return "Fair"


def build_valuation_df() -> pd.DataFrame:
    df_fr, df_c, df_s, df_mc_cur, df_mc_hist = load_data()

    # Merge base data
    df = df_c.merge(df_s, on="company_id", how="left")
    df = df.merge(df_mc_cur, on="company_id", how="left")
    df = df.merge(df_fr, on="company_id", how="left")

    # 5-year median P/E
    df_5yr = compute_five_yr_median_pe(df_mc_hist)
    df = df.merge(df_5yr, on="company_id", how="left")

    # FCF yield
    df["fcf_yield_pct"] = np.where(
        (df["market_cap_crore"].notna()) & (df["market_cap_crore"] > 0),
        df["free_cash_flow_cr"] / df["market_cap_crore"] * 100,
        np.nan
    )

    # Sector median P/E
    df_sec_pe = compute_sector_median_pe(df)
    df = df.merge(df_sec_pe, on="broad_sector", how="left")

    # P/E vs sector median %
    df["pe_vs_sector_median_pct"] = np.where(
        df["sector_median_pe"].notna() & (df["sector_median_pe"] != 0),
        (df["pe_ratio"] / df["sector_median_pe"] - 1) * 100,
        np.nan
    )

    # Flags
    df["flag"] = df.apply(lambda r: apply_valuation_flag(r["pe_ratio"], r["sector_median_pe"]), axis=1)

    # Rename and select output columns
    df = df.rename(columns={"broad_sector": "sector", "ev_ebitda": "ev_ebitda"})
    out_cols = ["company_id", "company_name", "sector", "pe_ratio", "pb_ratio", "ev_ebitda",
                "fcf_yield_pct", "five_yr_median_pe", "pe_vs_sector_median_pct", "sector_median_pe", "flag"]
    out_cols_present = [c for c in out_cols if c in df.columns]
    df_out = df[out_cols_present].copy()

    # Round numerics
    for col in df_out.select_dtypes(include="number").columns:
        df_out[col] = df_out[col].round(2)

    # Sort: Caution first, Discount second, Fair third, N/A last
    flag_order = {"Caution": 0, "Discount": 1, "Fair": 2, "N/A": 3}
    df_out["_sort"] = df_out["flag"].map(flag_order).fillna(9)
    df_out = df_out.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

    logger.info(f"Valuation summary: {len(df_out)} rows. "
                f"Caution={len(df_out[df_out['flag']=='Caution'])}, "
                f"Discount={len(df_out[df_out['flag']=='Discount'])}, "
                f"Fair={len(df_out[df_out['flag']=='Fair'])}")
    return df_out


def write_valuation_excel(df: pd.DataFrame, path: str):
    os.makedirs(os.path.dirname(path) or "output", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation Summary"

    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    flag_fills = {"Caution": CAUTION_FILL, "Discount": DISCOUNT_FILL, "Fair": FAIR_FILL}

    for _, row in df.iterrows():
        ws.append([row.get(c) for c in headers])
        cur_row = ws.max_row
        fill = flag_fills.get(str(row.get("flag", "")))
        if fill:
            for cell in ws[cur_row]:
                cell.fill = fill

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 22)

    wb.save(path)
    logger.info(f"Saved {path}")


def run():
    os.makedirs("output", exist_ok=True)
    logger.info("Building valuation summary...")
    df = build_valuation_df()

    excel_path = "output/valuation_summary.xlsx"
    write_valuation_excel(df, excel_path)
    logger.info(f"valuation_summary.xlsx: {len(df)} rows")

    # CSV of flagged companies only
    df_flags = df[df["flag"].isin(["Caution", "Discount"])].copy()
    flags_path = "output/valuation_flags.csv"
    df_flags.to_csv(flags_path, index=False)
    logger.info(f"valuation_flags.csv: {len(df_flags)} rows (Caution + Discount)")

    print(f"[OK] valuation_summary.xlsx -> {len(df)} rows")
    print(f"[OK] valuation_flags.csv -> {len(df_flags)} rows")
    return df


if __name__ == "__main__":
    run()
