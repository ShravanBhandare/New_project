"""
src/analytics/peer.py
Peer group percentile ranking engine.
- Loads peer_groups from SQLite
- Computes PERCENT_RANK for 10 metrics within each of 11 peer groups
- Inverts D/E rank (lower D/E = higher rank)
- Writes results to peer_percentiles table
- Generates matplotlib radar charts per company in reports/radar_charts/
- Generates output/peer_comparison.xlsx with 11 sheets (one per peer group)
"""

import os
import sqlite3
import logging
import warnings

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

DB_PATH = "data/nifty100.db"
RADAR_DIR = "reports/radar_charts"
PEER_EXCEL = "output/peer_comparison.xlsx"
ACTIVE_YEAR = "2024-03"

METRICS = [
    "return_on_equity_pct",   # ROE
    "roce_percentage",         # ROCE
    "net_profit_margin_pct",  # NPM
    "debt_to_equity",         # D/E (inverted)
    "free_cash_flow_cr",      # FCF
    "pat_cagr_5yr",           # PAT CAGR 5yr
    "revenue_cagr_5yr",       # Revenue CAGR 5yr
    "eps_cagr_5yr",           # EPS CAGR 5yr
    "interest_coverage",      # ICR
    "asset_turnover",         # Asset Turnover
]
METRIC_LABELS = {
    "return_on_equity_pct": "ROE",
    "roce_percentage": "ROCE",
    "net_profit_margin_pct": "Net Profit Margin",
    "debt_to_equity": "D/E",
    "free_cash_flow_cr": "FCF",
    "pat_cagr_5yr": "PAT CAGR 5yr",
    "revenue_cagr_5yr": "Rev CAGR 5yr",
    "eps_cagr_5yr": "EPS CAGR 5yr",
    "interest_coverage": "Interest Coverage",
    "asset_turnover": "Asset Turnover",
}
INVERTED_METRICS = {"debt_to_equity"}   # lower is better

RADAR_AXES = [
    "return_on_equity_pct",
    "roce_percentage",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "composite_quality_score",
]
RADAR_LABELS = ["ROE", "ROCE", "NPM", "D/E\n(inverted)", "FCF\nScore", "PAT CAGR", "Rev CAGR", "Comp.Score"]


# ─────────────────────────────────────────
# DB helpers
# ─────────────────────────────────────────
def get_conn():
    return sqlite3.connect(DB_PATH)


def ensure_peer_percentiles_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS peer_percentiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id TEXT,
            peer_group_name TEXT,
            metric TEXT,
            value REAL,
            percentile_rank REAL,
            year TEXT
        )
    """)
    conn.commit()


# ─────────────────────────────────────────
# Data loading
# ─────────────────────────────────────────
def load_peer_groups(conn) -> pd.DataFrame:
    """Returns: company_id, peer_group_name"""
    return pd.read_sql("SELECT company_id, peer_group_name FROM peer_groups", conn)


def load_ratios(conn) -> pd.DataFrame:
    df_fr = pd.read_sql(
        "SELECT * FROM financial_ratios WHERE year = ?", conn, params=(ACTIVE_YEAR,)
    )
    # Also load companies
    df_c = pd.read_sql("SELECT id as company_id, company_name FROM companies", conn)
    df_s = pd.read_sql("SELECT company_id, broad_sector FROM sectors", conn)

    # Try fetching roce_percentage from financial_ratios if it exists, else add zeros
    if "roce_percentage" not in df_fr.columns:
        df_fr["roce_percentage"] = np.nan
    if "revenue_cagr_5yr" not in df_fr.columns:
        # alias from sales_cagr_5yr
        if "sales_cagr_5yr" in df_fr.columns:
            df_fr["revenue_cagr_5yr"] = df_fr["sales_cagr_5yr"]
        else:
            df_fr["revenue_cagr_5yr"] = np.nan

    df = df_fr.merge(df_c, on="company_id", how="left")
    df = df.merge(df_s, on="company_id", how="left")
    return df


# ─────────────────────────────────────────
# Percentile computation
# ─────────────────────────────────────────
def compute_percent_rank(series: pd.Series) -> pd.Series:
    """Compute 0..1 percentile rank within a series (ignoring NaN)."""
    return series.rank(method="average", pct=True, na_option="keep")


def compute_peer_percentiles(df_ratios: pd.DataFrame, df_peers: pd.DataFrame) -> pd.DataFrame:
    merged = df_peers.merge(df_ratios, on="company_id", how="left")
    records = []

    for group_name, grp in merged.groupby("peer_group_name"):
        for metric in METRICS:
            if metric not in grp.columns:
                continue
            s = grp[metric].astype(float)
            pct = compute_percent_rank(s)
            if metric in INVERTED_METRICS:
                pct = 1 - pct   # invert so lower D/E ranks higher

            for idx, row in grp.iterrows():
                records.append({
                    "company_id": row["company_id"],
                    "peer_group_name": group_name,
                    "metric": metric,
                    "value": s[idx],
                    "percentile_rank": pct[idx],
                    "year": ACTIVE_YEAR,
                })

    return pd.DataFrame(records)


def write_percentiles_to_db(conn, df_pct: pd.DataFrame):
    conn.execute("DELETE FROM peer_percentiles WHERE year = ?", (ACTIVE_YEAR,))
    conn.commit()
    df_pct.to_sql("peer_percentiles", conn, if_exists="append", index=False)
    conn.commit()
    logger.info(f"Written {len(df_pct)} peer percentile rows to DB.")


# ─────────────────────────────────────────
# Normalise metric to 0-100 for radar
# ─────────────────────────────────────────
def winsorise_scale(series: pd.Series) -> pd.Series:
    """Winsorise to P10/P90 then scale to 0-100."""
    p10 = series.quantile(0.10)
    p90 = series.quantile(0.90)
    clipped = series.clip(lower=p10, upper=p90)
    rng = p90 - p10
    if rng == 0:
        return pd.Series([50.0] * len(series), index=series.index)
    return ((clipped - p10) / rng) * 100.0


def normalise_all_metrics(df_ratios: pd.DataFrame) -> pd.DataFrame:
    """Return a copy of df_ratios with each radar-axis metric normalised 0-100."""
    df = df_ratios.copy()
    for col in RADAR_AXES:
        if col in df.columns:
            norm = winsorise_scale(df[col].fillna(0))
            if col in INVERTED_METRICS:
                norm = 100 - norm
            df[col + "_norm"] = norm
    return df


# ─────────────────────────────────────────
# Radar chart generation
# ─────────────────────────────────────────
def make_radar_chart(company_id: str, company_name: str,
                     company_vals: list, avg_vals: list,
                     labels: list, peer_group: str, outpath: str):
    N = len(labels)
    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles += angles[:1]

    cv = company_vals + [company_vals[0]]
    av = avg_vals + [avg_vals[0]]

    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)

    ax.fill(angles, cv, alpha=0.25, color='#2196F3')
    ax.plot(angles, cv, linewidth=2, color='#2196F3', label=company_id)

    ax.plot(angles, av, linewidth=1.5, linestyle='--', color='#FF5722', label=f"{peer_group} Avg")

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_ylim(0, 100)
    ax.set_yticks([20, 40, 60, 80, 100])
    ax.set_yticklabels(['20', '40', '60', '80', '100'], fontsize=7, color='gray')
    ax.grid(color='gray', linestyle='--', linewidth=0.5, alpha=0.5)

    ax.set_title(f"{company_name}\nvs {peer_group} Average", fontsize=11, fontweight='bold', pad=20)
    ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=8)

    plt.tight_layout()
    plt.savefig(outpath, dpi=120, bbox_inches='tight')
    plt.close(fig)


def make_standalone_chart(company_id: str, company_name: str,
                           company_vals: list, nifty_avg_vals: list,
                           labels: list, outpath: str):
    N = len(labels)
    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles += angles[:1]

    cv = company_vals + [company_vals[0]]
    av = nifty_avg_vals + [nifty_avg_vals[0]]

    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    ax.set_theta_offset(np.pi / 2)
    ax.set_theta_direction(-1)

    ax.fill(angles, cv, alpha=0.25, color='#4CAF50')
    ax.plot(angles, cv, linewidth=2, color='#4CAF50', label=company_id)
    ax.plot(angles, av, linewidth=1.5, linestyle='--', color='#9C27B0', label='Nifty 100 Avg')

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_ylim(0, 100)
    ax.set_yticks([20, 40, 60, 80, 100])
    ax.set_yticklabels(['20', '40', '60', '80', '100'], fontsize=7, color='gray')
    ax.grid(color='gray', linestyle='--', linewidth=0.5, alpha=0.5)

    ax.set_title(f"{company_name}\nvs Nifty 100 Average", fontsize=11, fontweight='bold', pad=20)
    ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=8)

    plt.tight_layout()
    plt.savefig(outpath, dpi=120, bbox_inches='tight')
    plt.close(fig)


def generate_all_radar_charts(df_norm: pd.DataFrame, df_peers: pd.DataFrame):
    os.makedirs(RADAR_DIR, exist_ok=True)
    norm_cols = [c + "_norm" for c in RADAR_AXES if c + "_norm" in df_norm.columns]
    actual_axes = [c for c in RADAR_AXES if c + "_norm" in df_norm.columns]
    radar_labels = [RADAR_LABELS[RADAR_AXES.index(c)] for c in actual_axes]

    # Nifty 100 avg (for standalone)
    nifty_avg = [df_norm[c + "_norm"].mean() for c in actual_axes]

    peer_map = dict(zip(df_peers["company_id"], df_peers["peer_group_name"]))

    # Compute group averages
    df_with_peers = df_norm.merge(df_peers, on="company_id", how="left")
    group_avgs = {}
    for grp_name, grp in df_with_peers.groupby("peer_group_name"):
        group_avgs[grp_name] = [grp[c + "_norm"].mean() for c in actual_axes]

    generated = 0
    for _, row in df_norm.iterrows():
        cid = row["company_id"]
        cname = row.get("company_name", cid)
        outpath = os.path.join(RADAR_DIR, f"{cid}_radar.png")
        company_vals = [float(row.get(c + "_norm", 50.0)) for c in actual_axes]

        group = peer_map.get(cid)
        if group and group in group_avgs:
            make_radar_chart(cid, cname, company_vals, group_avgs[group],
                             radar_labels, group, outpath)
        else:
            make_standalone_chart(cid, cname, company_vals, nifty_avg, radar_labels, outpath)
        generated += 1

    logger.info(f"Generated {generated} radar charts in {RADAR_DIR}.")


# ─────────────────────────────────────────
# Peer Comparison Excel
# ─────────────────────────────────────────
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD_FILL   = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

DISPLAY_METRICS = METRICS  # 10 core metrics
PCT_COLS = [m + "_pct_rank" for m in DISPLAY_METRICS]


def build_peer_sheet_df(group_name: str, df_ratios: pd.DataFrame,
                         df_peers: pd.DataFrame, df_pct: pd.DataFrame) -> pd.DataFrame:
    members = df_peers[df_peers["peer_group_name"] == group_name]["company_id"].tolist()
    df_sub = df_ratios[df_ratios["company_id"].isin(members)].copy()

    # Add percentile ranks
    for metric in DISPLAY_METRICS:
        sub_pct = df_pct[(df_pct["peer_group_name"] == group_name) & (df_pct["metric"] == metric)]
        pct_map = dict(zip(sub_pct["company_id"], sub_pct["percentile_rank"]))
        df_sub[metric + "_pct_rank"] = df_sub["company_id"].map(pct_map)

    cols_order = ["company_id", "company_name"] + DISPLAY_METRICS + PCT_COLS
    existing_cols = [c for c in cols_order if c in df_sub.columns]
    return df_sub[existing_cols].copy()


def write_peer_comparison_excel(df_ratios: pd.DataFrame, df_peers: pd.DataFrame, df_pct: pd.DataFrame):
    os.makedirs(os.path.dirname(PEER_EXCEL), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    peer_groups = sorted(df_peers["peer_group_name"].unique())
    for group_name in peer_groups:
        ws = wb.create_sheet(title=group_name[:31])

        df_sheet = build_peer_sheet_df(group_name, df_ratios, df_peers, df_pct)
        if df_sheet.empty:
            ws.append([f"No data for {group_name}"])
            continue

        # Compute benchmark: row with highest composite_quality_score
        benchmark_id = None
        if "composite_quality_score" in df_sheet.columns:
            idx_best = df_sheet["composite_quality_score"].idxmax()
            benchmark_id = df_sheet.loc[idx_best, "company_id"]

        # Compute median summary
        numeric_cols = df_sheet.select_dtypes(include=[np.number]).columns.tolist()
        median_row = df_sheet[numeric_cols].median().to_dict()

        # Write headers
        headers = list(df_sheet.columns)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for _, row in df_sheet.iterrows():
            ws.append(list(row))
            cur_row = ws.max_row
            is_benchmark = row.get("company_id") == benchmark_id

            # Highlight benchmark in gold
            if is_benchmark:
                for cell in ws[cur_row]:
                    cell.fill = GOLD_FILL

            # Colour-code percentile rank columns
            for col_idx, col_name in enumerate(headers, start=1):
                if col_name.endswith("_pct_rank"):
                    val = row.get(col_name)
                    if val is None or (isinstance(val, float) and np.isnan(val)):
                        continue
                    cell = ws.cell(row=cur_row, column=col_idx)
                    if not is_benchmark:   # don't override gold
                        if val >= 0.75:
                            cell.fill = GREEN_FILL
                        elif val >= 0.25:
                            cell.fill = YELLOW_FILL
                        else:
                            cell.fill = RED_FILL

        # Median summary row
        median_vals = ["MEDIAN", "—"]
        for col_name in headers[2:]:
            median_vals.append(round(median_row.get(col_name, np.nan), 2))
        ws.append(median_vals)
        med_row = ws.max_row
        for cell in ws[med_row]:
            cell.font = Font(bold=True, italic=True)

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 22)

        logger.info(f"Sheet '{group_name}' written with {df_sheet.shape[0]} companies.")

    wb.save(PEER_EXCEL)
    logger.info(f"Peer comparison Excel saved to {PEER_EXCEL}")


# ─────────────────────────────────────────
# Main
# ─────────────────────────────────────────
def run():
    conn = get_conn()
    ensure_peer_percentiles_table(conn)

    logger.info("Loading peer groups and financial ratios...")
    df_peers = load_peer_groups(conn)
    df_ratios = load_ratios(conn)

    peer_group_count = df_peers["peer_group_name"].nunique()
    logger.info(f"Loaded {len(df_peers)} peer assignments across {peer_group_count} groups.")

    companies_with_no_group = set(df_ratios["company_id"]) - set(df_peers["company_id"])
    if companies_with_no_group:
        logger.info(f"{len(companies_with_no_group)} companies have no peer group assigned.")

    # Compute percentiles
    logger.info("Computing peer percentile ranks...")
    df_pct = compute_peer_percentiles(df_ratios, df_peers)
    write_percentiles_to_db(conn, df_pct)

    # Normalise metrics for radar
    logger.info("Generating radar charts...")
    df_norm = normalise_all_metrics(df_ratios)
    generate_all_radar_charts(df_norm, df_peers)

    # Excel peer comparison
    logger.info("Building peer comparison Excel...")
    write_peer_comparison_excel(df_ratios, df_peers, df_pct)

    conn.close()
    logger.info("Peer engine complete.")


if __name__ == "__main__":
    run()
