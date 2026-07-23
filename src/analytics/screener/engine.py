"""
src/analytics/screener/engine.py
Custom filter engine:
- Loads config/screener_config.yaml
- Builds master screener DataFrame for the active year
- Computes sector-relative composite score (Winsorised 0-100)
- Applies preset filters via pandas query
- Exports output/screener_output.xlsx with 20 KPI columns + colour-coded cells
"""

import sqlite3
import pandas as pd
import numpy as np
import yaml
import os
import logging
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

DB_PATH = "data/nifty100.db"
CONFIG_PATH = "config/screener_config.yaml"

# ─── Colour fills ───────────────────────────────────────────
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


# ─── Winsorise + scale helper ────────────────────────────────
def winsorise_scale(series: pd.Series, invert: bool = False) -> pd.Series:
    """Clip to P10/P90 and scale to 0-100. If invert=True, lower raw = higher score."""
    valid = series.dropna()
    if len(valid) < 2:
        return pd.Series([50.0] * len(series), index=series.index)
    p10 = valid.quantile(0.10)
    p90 = valid.quantile(0.90)
    clipped = series.clip(lower=p10, upper=p90)
    rng = p90 - p10
    if rng == 0:
        return pd.Series([50.0] * len(series), index=series.index)
    scaled = ((clipped - p10) / rng) * 100.0
    if invert:
        scaled = 100.0 - scaled
    return scaled


# ─── Composite score (35/30/20/15 weights) ───────────────────
def compute_composite(df: pd.DataFrame) -> pd.Series:
    """
    Composite Quality Score (0-100):
      35% Profitability: ROE 15% + ROCE 10% + NPM 10%
      30% Cash Quality:  FCF CAGR 15% (proxy: FCF/market_cap) + CFO/PAT 10% + FCF flag 5%
      20% Growth:        Rev CAGR 5yr 10% + PAT CAGR 5yr 10%
      15% Leverage:      D/E score 10% + ICR score 5%
    """
    def sc(col, invert=False):
        if col in df.columns:
            return winsorise_scale(df[col].fillna(0), invert=invert)
        return pd.Series([50.0] * len(df), index=df.index)

    roe_s  = sc("return_on_equity_pct")
    roce_s = sc("roce_percentage") if "roce_percentage" in df.columns else sc("return_on_equity_pct")
    npm_s  = sc("net_profit_margin_pct")

    fcf_flag_s = (df["free_cash_flow_cr"].fillna(0) > 0).astype(float) * 100.0
    cfo_pat_s  = sc("cfo_quality_score") if "cfo_quality_score" in df.columns else pd.Series([50.0]*len(df), index=df.index)
    fcf_scale_s = sc("free_cash_flow_cr")

    rev_cagr_col = "sales_cagr_5yr" if "revenue_cagr_5yr" not in df.columns else "revenue_cagr_5yr"
    rev_s  = sc(rev_cagr_col)
    pat_s  = sc("pat_cagr_5yr")

    de_s   = sc("debt_to_equity", invert=True)
    icr_s  = sc("interest_coverage")

    profitability = roe_s * 0.15 + roce_s * 0.10 + npm_s * 0.10
    cash_quality  = fcf_scale_s * 0.15 + cfo_pat_s * 0.10 + fcf_flag_s * 0.05
    growth        = rev_s * 0.10 + pat_s * 0.10
    leverage      = de_s * 0.10 + icr_s * 0.05

    # Scale each pillar to their max possible contribution
    raw = profitability + cash_quality + growth + leverage  # max ≈ 35+30+20+15 = 100
    return raw.clip(0, 100).round(2)


def compute_sector_relative(df: pd.DataFrame, score_col: str = "composite_quality_score") -> pd.Series:
    """Normalise composite score within each broad_sector (0-100 within peer sector)."""
    out = pd.Series(np.nan, index=df.index)
    for sector, grp in df.groupby("sector"):
        s = grp[score_col]
        mn, mx = s.min(), s.max()
        if mx == mn:
            out.loc[grp.index] = 50.0
        else:
            out.loc[grp.index] = ((s - mn) / (mx - mn) * 100).round(2)
    return out


# ─── Load screener data ──────────────────────────────────────
def get_screener_data(active_year: str = "2024-03") -> pd.DataFrame:
    if not os.path.exists(DB_PATH):
        logger.error(f"DB not found: {DB_PATH}")
        return pd.DataFrame()

    conn = sqlite3.connect(DB_PATH)

    df_comp = pd.read_sql("SELECT id as company_id, company_name FROM companies", conn)
    df_sec  = pd.read_sql("SELECT company_id, broad_sector as sector, sub_sector FROM sectors", conn)
    df_fr   = pd.read_sql("SELECT * FROM financial_ratios WHERE year = ?", conn, params=(active_year,))

    if df_fr.empty:
        logger.warning(f"No financial_ratios rows for year {active_year}")
        conn.close()
        return pd.DataFrame()

    # Previous year D/E for de_declining
    try:
        parts = active_year.split("-")
        prev_year = f"{int(parts[0]) - 1}-{parts[1]}"
    except Exception:
        prev_year = None

    if prev_year:
        df_prev = pd.read_sql(
            "SELECT company_id, debt_to_equity as prev_debt_to_equity FROM financial_ratios WHERE year = ?",
            conn, params=(prev_year,)
        )
    else:
        df_prev = pd.DataFrame(columns=["company_id", "prev_debt_to_equity"])

    try:
        yr_int = int(active_year.split("-")[0])
    except Exception:
        yr_int = 2024

    df_mc = pd.read_sql(
        "SELECT company_id, market_cap_crore, pe_ratio, pb_ratio, ev_ebitda, dividend_yield_pct "
        "FROM market_cap WHERE year = ?", conn, params=(yr_int,)
    )
    df_pl = pd.read_sql(
        "SELECT company_id, sales, net_profit FROM profitandloss WHERE year = ?",
        conn, params=(active_year,)
    )
    conn.close()

    # Merge
    df = df_fr.merge(df_comp, on="company_id", how="inner")
    df = df.merge(df_sec,  on="company_id", how="left")
    df = df.merge(df_mc,   on="company_id", how="left")
    df = df.merge(df_pl,   on="company_id", how="left", suffixes=("", "_pl"))
    df = df.merge(df_prev, on="company_id", how="left")

    # Fill sales from P&L if needed
    if "sales" not in df.columns:
        df["sales"] = df.get("sales_pl", np.nan)
    else:
        if "sales_pl" in df.columns:
            df["sales"] = df["sales"].fillna(df["sales_pl"])

    # Alias
    if "revenue_cagr_5yr" not in df.columns and "sales_cagr_5yr" in df.columns:
        df["revenue_cagr_5yr"] = df["sales_cagr_5yr"]

    # FCF yield
    def calc_fcf_yield(row):
        fcf = row.get("free_cash_flow_cr")
        mc  = row.get("market_cap_crore")
        if pd.isna(fcf) or pd.isna(mc) or mc <= 0:
            return 0.0
        return (fcf / mc) * 100.0

    df["fcf_yield"] = df.apply(calc_fcf_yield, axis=1)

    # D/E declining flag
    def check_de_declining(row):
        de = row.get("debt_to_equity")
        prev = row.get("prev_debt_to_equity")
        if pd.isna(de) or pd.isna(prev):
            return False
        return bool(de < prev)

    df["de_declining"] = df.apply(check_de_declining, axis=1)

    # Round D/E to 2dp (so virtually-zero shows as 0.00)
    if "debt_to_equity" in df.columns:
        df["debt_to_equity"] = df["debt_to_equity"].round(2)

    # Composite scores
    df["composite_quality_score"] = compute_composite(df)
    df["sector_relative_score"]   = compute_sector_relative(df, "composite_quality_score")

    return df


# ─── Preset filter ───────────────────────────────────────────
COL_MAP = {
    "return_on_equity_pct": "return_on_equity_pct",
    "debt_to_equity":       "debt_to_equity",
    "free_cash_flow_cr":    "free_cash_flow_cr",
    "sales_cagr_5yr":       "sales_cagr_5yr",
    "sales_cagr_3yr":       "sales_cagr_3yr",
    "revenue_cagr_5yr":     "revenue_cagr_5yr",
    "pe_ratio":             "pe_ratio",
    "pb_ratio":             "pb_ratio",
    "dividend_yield_pct":   "dividend_yield_pct",
    "dividend_payout_ratio_pct": "dividend_payout_ratio_pct",
    "pat_cagr_5yr":         "pat_cagr_5yr",
    "sales":                "sales",
    "composite_score":      "composite_quality_score",
    "de_declining":         "de_declining",
    "net_profit_margin_pct": "net_profit_margin_pct",
    "interest_coverage":    "interest_coverage",
    "asset_turnover":       "asset_turnover",
}


def apply_preset_filters(df: pd.DataFrame, preset_name: str, config: dict) -> pd.DataFrame:
    preset = config["presets"].get(preset_name)
    if not preset:
        logger.error(f"Preset {preset_name} not found.")
        return pd.DataFrame()

    filtered = df.copy()
    query_parts = []

    for rule in preset["filters"]:
        col = COL_MAP.get(rule["metric"], rule["metric"])
        op  = rule["operator"]
        val = rule["value"]
        if col not in filtered.columns:
            logger.warning(f"Column {col} missing — skipping filter.")
            continue
        if op == "==":
            query_parts.append(f"{col} == {val}")
        else:
            query_parts.append(f"{col} {op} {val}")

    if query_parts:
        q = " and ".join(query_parts)
        logger.info(f"Preset '{preset_name}': {q}")
        try:
            filtered = filtered.query(q)
        except Exception as e:
            logger.error(f"Query failed '{q}': {e}")
            # Manual fallback
            for rule in preset["filters"]:
                col = COL_MAP.get(rule["metric"], rule["metric"])
                op, val = rule["operator"], rule["value"]
                if col not in filtered.columns:
                    continue
                ops = {">": lambda s,v: s>v, "<": lambda s,v: s<v, "==": lambda s,v: s==v,
                       ">=": lambda s,v: s>=v, "<=": lambda s,v: s<=v}
                if op in ops:
                    filtered = filtered[ops[op](filtered[col], val)]

    rank_col = COL_MAP.get(preset.get("ranking_metric", "composite_quality_score"),
                           preset.get("ranking_metric", "composite_quality_score"))
    ascending = preset.get("sort_order", "desc") == "asc"
    if rank_col in filtered.columns:
        filtered = filtered.sort_values(by=rank_col, ascending=ascending)

    return filtered


# ─── 20 KPI display columns ──────────────────────────────────
DISPLAY_COLS = [
    "company_id", "company_name", "sector",
    "return_on_equity_pct", "roce_percentage", "net_profit_margin_pct",
    "debt_to_equity", "interest_coverage", "asset_turnover",
    "free_cash_flow_cr", "capex_cr", "cash_from_operations_cr",
    "earnings_per_share", "book_value_per_share",
    "dividend_payout_ratio_pct", "dividend_yield_pct",
    "sales_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "composite_quality_score", "sector_relative_score",
    "pe_ratio", "pb_ratio",
]


# ─── Colour-code helper ──────────────────────────────────────
def colour_code_sheet(ws, df_filtered: pd.DataFrame, preset_rules: list, headers: list):
    """Apply green/red fills based on whether each cell passes its preset threshold."""
    rule_map = {}  # col_name -> (operator, value)
    for rule in preset_rules:
        col = COL_MAP.get(rule["metric"], rule["metric"])
        rule_map[col] = (rule["operator"], rule["value"])

    # Row offset: row 1 = header, data starts row 2
    for r_idx, (_, row) in enumerate(df_filtered.iterrows(), start=2):
        for c_idx, col_name in enumerate(headers, start=1):
            if col_name not in rule_map:
                continue
            op, val = rule_map[col_name]
            cell_val = row.get(col_name)
            if cell_val is None or (isinstance(cell_val, float) and np.isnan(cell_val)):
                continue
            try:
                passes = eval(f"{cell_val} {op} {val}")
            except Exception:
                passes = False
            ws.cell(row=r_idx, column=c_idx).fill = GREEN_FILL if passes else RED_FILL


# ─── Export Excel ────────────────────────────────────────────
def run_screener_and_export(active_year: str = "2024-03"):
    logger.info(f"Running custom screener engine for {active_year}…")

    if not os.path.exists(CONFIG_PATH):
        logger.error(f"Config not found: {CONFIG_PATH}")
        return

    with open(CONFIG_PATH) as f:
        config = yaml.safe_load(f)

    df = get_screener_data(active_year)
    if df.empty:
        logger.error("No data loaded.")
        return

    logger.info(f"Loaded {len(df)} records for {active_year}.")
    os.makedirs("output", exist_ok=True)
    output_path = "output/screener_output.xlsx"

    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for preset_name, preset_cfg in config["presets"].items():
        preset_df = apply_preset_filters(df, preset_name, config)
        sheet_label = preset_cfg.get("label", preset_name.replace("_", " ").title())
        ws = wb.create_sheet(title=sheet_label[:31])

        cols_to_use = [c for c in DISPLAY_COLS if c in preset_df.columns]
        out_df = preset_df[cols_to_use].reset_index(drop=True)

        # Round numerics
        for col in out_df.select_dtypes(include=[np.number]).columns:
            out_df[col] = out_df[col].round(2)

        # Write header
        ws.append(cols_to_use)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for _, row in out_df.iterrows():
            ws.append([row.get(c) for c in cols_to_use])

        # Colour-code
        colour_code_sheet(ws, out_df, preset_cfg.get("filters", []), cols_to_use)

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 24)

        logger.info(f"Sheet '{sheet_label}' → {len(out_df)} companies.")

    wb.save(output_path)
    logger.info(f"Saved screener output to {output_path}")


if __name__ == "__main__":
    run_screener_and_export()
